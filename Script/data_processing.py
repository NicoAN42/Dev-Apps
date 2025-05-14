import os
import math
import time
import threading
import tkinter as tk
from datetime import datetime
from tkinter import messagebox, filedialog, scrolledtext
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.toast import ToastNotification

class DataProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🚀 Processing Data Mission Digital")
        self.center_window(760, 560)

        self.style = tb.Style("minty")
        self.current_theme = "minty"

        self.frame = tb.Frame(self.root, padding=30)
        self.frame.pack(expand=True, fill="both")

        # Title
        self.title = tb.Label(self.frame, text="Processing Data Mission Digital",
                              font=("Segoe UI", 20, "bold"), bootstyle="primary")
        self.title.grid(row=0, column=0, pady=(0, 30))

        # Center content
        content = tb.Frame(self.frame)
        content.grid(row=1, column=0)
        content.columnconfigure(0, weight=1)
        content.columnconfigure(1, weight=1)

        # Form Inputs
        tb.Label(content, text="Nama:", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, sticky="e", padx=10, pady=5)
        self.nama_entry = tb.Entry(content, width=24)
        self.nama_entry.grid(row=0, column=1, pady=5)

        tb.Label(content, text="NIP:", font=("Segoe UI", 12, "bold")).grid(row=1, column=0, sticky="e", padx=10, pady=5)
        self.nip_entry = tb.Entry(content, width=24)
        self.nip_entry.grid(row=1, column=1, pady=5)

        # Button: Select + Start
        self.start_button = tb.Button(content, text="📁 Choose Folder & ▶ Start",
                                      command=self.select_and_start, bootstyle="success-outline", width=35)
        self.start_button.grid(row=2, column=0, columnspan=2, pady=(15, 5))

        # Dark Mode Toggle
        self.theme_toggle = tb.Checkbutton(content, text="🌗 Dark Mode", bootstyle="info-round-toggle",
                                           command=self.toggle_theme)
        self.theme_toggle.grid(row=3, column=0, columnspan=2, pady=(10, 20))

        # Progress
        self.progress = tb.Progressbar(self.frame, length=680, bootstyle="success-striped")
        self.progress.grid(row=2, column=0, pady=15)

        # Status & Time
        status_frame = tb.Frame(self.frame)
        status_frame.grid(row=3, column=0, pady=5)
        self.status_label = tb.Label(status_frame, text="Status: Waiting", font=("Segoe UI", 10))
        self.status_label.pack(side="left", padx=10)
        self.elapsed_label = tb.Label(status_frame, text="Elapsed time: 0s", font=("Segoe UI", 10))
        self.elapsed_label.pack(side="right", padx=10)

        # Log Box
        self.log_box = scrolledtext.ScrolledText(self.frame, height=8, font=("Consolas", 10))
        self.log_box.grid(row=4, column=0, padx=10, pady=(10, 15), sticky="nsew")
        self.log_box.config(state="disabled", borderwidth=2, relief="groove")

        # Cancel Button
        self.cancel_button = tb.Button(self.frame, text="⛔ Cancel", command=self.cancel_process,
                                       state="disabled", bootstyle="danger")
        self.cancel_button.grid(row=5, column=0, pady=5)

        # Variables
        self.folder_path = None
        self.stop_flag = False
        self.start_time = None
        self.update_timer = None
        self.log_path = None

    def center_window(self, w, h):
        self.root.update_idletasks()
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        x = (screen_w // 2) - (w // 2)
        y = (screen_h // 2) - (h // 2)
        self.root.geometry(f"{w}x{h}+{x}+{y}")

    def toggle_theme(self):
        if self.current_theme == "minty":
            self.style.theme_use("darkly")
            self.current_theme = "darkly"
        else:
            self.style.theme_use("minty")
            self.current_theme = "minty"

    def log(self, message):
        self.log_box.config(state="normal")
        self.log_box.insert(tk.END, message + "\n")
        self.log_box.see(tk.END)
        self.log_box.config(state="disabled")
        if self.log_path:
            with open(self.log_path, "a", encoding="utf-8") as f:
                f.write(message + "\n")

    def select_and_start(self):
        self.folder_path = filedialog.askdirectory()
        if self.folder_path:
            self.status_label.config(text=f"Selected: {self.folder_path}")
            self.log(f"📂 Folder selected: {self.folder_path}")
            self.start_process()

    def start_process(self):
        nama = self.nama_entry.get().strip()
        nip = self.nip_entry.get().strip()
        if not nama or not nip:
            messagebox.showwarning("Input Required", "Please enter both Nama and NIP.")
            return
        if not self.folder_path:
            messagebox.showwarning("No Folder", "Please select a parent folder.")
            return

        # Disable UI
        self.nama_entry.config(state="disabled")
        self.nip_entry.config(state="disabled")
        self.start_button.config(state="disabled")
        self.cancel_button.config(state="normal")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.log_path = os.path.join(self.folder_path, f"log_processing_{timestamp}.txt")
        with open(self.log_path, "w", encoding="utf-8") as f:
            f.write(f"Log started: {datetime.now()}\n")

        self.stop_flag = False
        self.start_time = time.time()
        self.update_elapsed_time()
        threading.Thread(target=self.run_processing, args=(nama, nip)).start()

    def cancel_process(self):
        self.stop_flag = True
        self.status_label.config(text="Cancelling...")
        self.log("⚠️ Cancel requested by user.")

    def update_elapsed_time(self):
        if self.start_time:
            elapsed = int(time.time() - self.start_time)
            mins, secs = divmod(elapsed, 60)
            self.elapsed_label.config(text=f"Elapsed time: {mins}m {secs}s" if mins else f"Elapsed time: {secs}s")
            self.update_timer = self.root.after(1000, self.update_elapsed_time)

    def run_processing(self, nama, nip):
        created_count = 0
        skipped = []
        all_dirs = [dirpath for dirpath, _, _ in os.walk(self.folder_path)]
        total = len(all_dirs)
        self.progress.config(maximum=total, value=0)

        for i, dirpath in enumerate(all_dirs, 1):
            if self.stop_flag:
                break
            combined = os.path.join(dirpath, "combined.xlsx")
            result = os.path.join(dirpath, "result.xlsx")

            if os.path.exists(combined):
                success, msg = self.calculate_and_save_excel(combined, result, nama, nip)
                if success:
                    created_count += 1
                    self.log(f"✅ Processed: {dirpath}")
                else:
                    self.log(f"❌ Failed: {dirpath} - {msg}")
            else:
                skipped.append(dirpath)
                self.log(f"⏭️ Skipped (no combined.xlsx): {dirpath}")

            self.progress["value"] = i
            self.status_label.config(text=f"Processing ({i}/{total})")
            self.root.update_idletasks()

        if self.update_timer:
            self.root.after_cancel(self.update_timer)

        elapsed = int(time.time() - self.start_time)
        mins, secs = divmod(elapsed, 60)
        time_str = f"{mins}m {secs}s" if mins else f"{secs}s"

        self.log("\n📋 SUMMARY:")
        self.log(f"Total folders: {total}")
        self.log(f"Files created: {created_count}")
        self.log(f"Skipped: {len(skipped)}")
        for d in skipped:
            self.log(f" - {d}")

        self.status_label.config(text=f"Done. Files created: {created_count}")
        self.elapsed_label.config(text=f"Elapsed time: {time_str}")

        self.cancel_button.config(state="disabled")
        self.nama_entry.config(state="normal")
        self.nip_entry.config(state="normal")
        self.start_button.config(state="normal")

        # Auto-close confirmation
        toast = ToastNotification(title="Finished",
                                  message="✅ Processing completed successfully.",
                                  duration=2000,
                                  bootstyle="success")
        toast.show_toast()
        
         # Auto-close confirmation dialog
        answer = messagebox.askyesno("Process Completed", "✅ Processing completed successfully.\n\nDo you want to close the application?")
        if answer:
            self.root.destroy()

    def calculate_and_save_excel(self, file_path, save_path, nama, nip):
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            bold = Font(bold=True)
            red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

            calc_ws = wb.create_sheet("Calculated Data")
            calc_ws.append(["Periode", "Mesin", "Counter", "Total", "Digital %"])
            for cell in calc_ws[1]: cell.font = bold; cell.border = border

            total_vals, pct_vals = [], []

            for row in ws.iter_rows(min_row=2, values_only=True):
                periode, mesin, counter = row
                total = mesin + counter
                pct = mesin / total if total else 0
                total_vals.append(total)
                pct_vals.append(pct)
                calc_ws.append([periode, mesin, counter, total, f"{pct*100:.2f}%".replace('.', ',')])

            for row in calc_ws.iter_rows(min_row=2, max_col=5):
                for c in row: c.border = border
                try:
                    val = float(row[4].value.replace('%', '').replace(',', '.'))
                    if val < 85: row[4].fill = red
                except: pass

            result_ws = wb.create_sheet("Result Data")
            result_ws.append(["Keterangan", "Presentase", "Transaksi"])
            for c in result_ws[1]: c.font = bold; c.border = border

            avg_pct = sum(pct_vals)/len(pct_vals)
            avg_trx = sum(total_vals)/len(total_vals)
            gap = 0.85 - avg_pct
            need_trx = int(gap * sum(total_vals))

            def pct(v): return f"{v*100:.2f}%".replace('.', ',')
            def num(v): return f"{round(v):,}".replace(',', '.')

            now_str = datetime.now().strftime("%d %B %Y")
            result_ws.append([f"Presentase per {now_str}", pct(avg_pct), ""])
            result_ws.append(["Target", pct(0.85), ""])
            result_ws.append(["Kekurangan transaksi", pct(gap), str(need_trx)])
            result_ws.append(["Rata-rata transaksi harian", "", str(round(avg_trx))])

            for row in result_ws.iter_rows(min_row=2, max_col=3):
                for c in row: c.border = border

            sim_ws = wb.create_sheet("Simulasi Transaksi")
            sim_ws.append(["Total Hari", "Target Transaksi"])
            for c in sim_ws[1]: c.font = bold; c.border = border

            for days in [10, 20, 30, 60, 90, 120, 180, 240]:
                sim_trx = need_trx + math.ceil(avg_trx / days)
                sim_ws.append([days, num(sim_trx)])

            for row in sim_ws.iter_rows(min_row=2, max_col=2):
                for c in row: c.border = border

            wb.save(save_path)
            return True, "Saved"
        except Exception as e:
            return False, str(e)

if __name__ == "__main__":
    root = tb.Window(themename="minty")
    app = DataProcessorApp(root)
    root.mainloop()
