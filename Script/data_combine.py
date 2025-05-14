import os
import threading
import time
from datetime import datetime, timedelta
from tkinter import Text, Scrollbar, StringVar, END, NORMAL, DISABLED, filedialog, messagebox
from ttkbootstrap import Window, Frame, Label, Entry, Button
from ttkbootstrap.constants import *
from ttkbootstrap.toast import ToastNotification
from ttkbootstrap.widgets import Progressbar, Checkbutton
from openpyxl import load_workbook, Workbook
from copy import copy

stop_flag = False
log_messages = []

def copy_sheet(src_ws, dest_wb, sheet_name):
    dest_ws = dest_wb.create_sheet(title=sheet_name)
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dest_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for col in src_ws.column_dimensions:
        dest_ws.column_dimensions[col].width = src_ws.column_dimensions[col].width
    for row in src_ws.row_dimensions:
        dest_ws.row_dimensions[row].height = src_ws.row_dimensions[row].height

def log_to_interface(message):
    global log_messages
    log_messages.append(message)
    log_text.config(state=NORMAL)
    log_text.insert(END, message + "\n")
    log_text.see(END)
    log_text.config(state=DISABLED)

def save_log_file(parent_folder):
    if not log_messages:
        return
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    log_filename = f"log_combined_{timestamp}.txt"
    log_path = os.path.join(parent_folder, log_filename)
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("\n".join(log_messages))
        log_to_interface(f"[LOG SAVED] {log_path}")
    except Exception as e:
        log_to_interface(f"[LOG ERROR] Failed to save log file: {e}")

def process_folder(folder_path, name, nip):
    trx_file = os.path.join(folder_path, "Trx.xlsx")
    act_file = os.path.join(folder_path, "Act.xlsx")
    output_file = os.path.join(folder_path, "combined.xlsx")

    if not os.path.isfile(trx_file) or not os.path.isfile(act_file):
        return False

    try:
        wb_trx = load_workbook(trx_file)
        wb_act = load_workbook(act_file)
        ws_trx = wb_trx["Sheet1"]
        ws_act = wb_act["MyWorkSheet-1"]

        new_wb = Workbook()
        new_wb.remove(new_wb.active)

        copy_sheet(ws_trx, new_wb, "Daily Transactions")
        copy_sheet(ws_act, new_wb, "Transaction Details")

        info_ws = new_wb.create_sheet("User Info")
        info_ws.append(["Name", name])
        info_ws.append(["NIP", nip])
        info_ws.append(["Date Created", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        new_wb.save(output_file)
        log_to_interface(f"[SUCCESS] {folder_path} -> combined.xlsx created.")
        return True
    except Exception as e:
        log_to_interface(f"[ERROR] {folder_path} -> {str(e)}")
        return False

def choose_and_process():
    global stop_flag
    stop_flag = False

    name = name_var.get().strip()
    nip = nip_var.get().strip()

    if not name or not nip:
        messagebox.showwarning("Input Error", "Name and NIP are required.")
        return

    folder = filedialog.askdirectory(title="Select Parent Folder")
    if not folder:
        return

    folder_var.set(folder)
    name_entry.config(state=DISABLED)
    nip_entry.config(state=DISABLED)
    process_button.config(state=DISABLED)
    cancel_button.config(state=NORMAL)

    log_text.config(state=NORMAL)
    log_text.delete("1.0", END)
    log_text.config(state=DISABLED)

    threading.Thread(target=run_processing, args=(name, nip, folder)).start()

def cancel_process():
    global stop_flag
    stop_flag = True
    progress_label.config(text="❌ Process canceled by user.")

def run_processing(name, nip, parent_folder):
    global stop_flag
    start_time = time.time()
    progress_label.config(text="🔍 Scanning folders...")
    folders = []
    missing_folders = []

    for dirpath, _, filenames in os.walk(parent_folder):
        if "Trx.xlsx" in filenames and "Act.xlsx" in filenames:
            folders.append(dirpath)
        elif "Trx.xlsx" not in filenames or "Act.xlsx" not in filenames:
            missing_folders.append(dirpath)

    total = len(folders)
    if total == 0:
        progress_label.config(text="⚠️ No valid folders found.")
        return

    progress_bar["maximum"] = total
    success_count = 0

    for i, folder in enumerate(folders, start=1):
        if stop_flag:
            break

        elapsed = time.time() - start_time
        avg_time = elapsed / i
        remaining = avg_time * (total - i)
        eta_time = datetime.now() + timedelta(seconds=remaining)

        if process_folder(folder, name, nip):
            success_count += 1

        progress_bar["value"] = i
        progress_label.config(text=f"✅ Processing folder {i}/{total}")
        elapsed_label.config(text=f"⏱️ Elapsed: {int(elapsed)}s")
        eta_label.config(text=f"⏳ ETA: {int(remaining)}s")
        finish_label.config(text=f"🕓 Estimated Finish: {eta_time.strftime('%H:%M:%S')}")
        root.update_idletasks()

    for folder in missing_folders:
        log_to_interface(f"[SKIPPED] {folder}")

    log_to_interface("")
    log_to_interface(f"Total folders processed: {total}")
    log_to_interface(f"Total files created: {success_count}")
    log_to_interface(f"Folders skipped: {len(missing_folders)}")

    save_log_file(parent_folder)

    cancel_button.config(state=DISABLED)
    elapsed = int(time.time() - start_time)
    if stop_flag:
        progress_label.config(text=f"❌ Process canceled. {success_count} out of {total} completed.")
    else:
        progress_label.config(text=f"🎉 DONE: {success_count}/{total} folders successfully processed.")
        messagebox.showinfo("Done", f"Processing complete!\nTotal time: {elapsed} seconds")
        show_toast(f"{success_count}/{total} folders processed successfully!")
        root.after(1500, root.destroy)

def toggle_theme():
    selected_theme = "darkly" if darkmode_var.get() else "minty"
    root.style.theme_use(selected_theme)

def show_toast(message):
    toast = ToastNotification(
        title="Process Complete",
        message=message,
        duration=5000,
        bootstyle="success",
        position=(root.winfo_rootx() + root.winfo_width() - 300, root.winfo_rooty() + root.winfo_height() - 100)
    )
    toast.show_toast()

# --- GUI ---
root = Window(themename="minty")
root.title("📁 Combine Data")
root.geometry("900x650")
root.resizable(False, False)
root.style.theme_use("minty")

window_width = 900
window_height = 650
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = int((screen_width - window_width) / 2)
y = int((screen_height - window_height) / 2)
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

Label(root, text="Combine Data", font=("Segoe UI", 18, "bold")).pack(pady=15)

form = Frame(root, padding=10)
form.pack()

Label(form, text="Name:", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, padx=10, pady=8, sticky='e')
name_var = StringVar()
name_entry = Entry(form, textvariable=name_var, width=15, font=("Segoe UI", 11))
name_entry.grid(row=0, column=1, padx=10, pady=8)

Label(form, text="NIP:", font=("Segoe UI", 11, "bold")).grid(row=1, column=0, padx=10, pady=8, sticky='e')
nip_var = StringVar()
nip_entry = Entry(form, textvariable=nip_var, width=15, font=("Segoe UI", 11))
nip_entry.grid(row=1, column=1, padx=10, pady=8)

folder_var = StringVar()

process_button = Button(root, text="📂 Select Folder & Start", bootstyle="success", width=30, command=choose_and_process)
process_button.pack(pady=10)

cancel_button = Button(root, text="❌ Cancel", bootstyle="danger", width=30, command=cancel_process, state=DISABLED)
cancel_button.pack(pady=5)

progress_bar = Progressbar(root, orient="horizontal", length=700, mode="determinate", style="success.Horizontal.TProgressbar")
progress_bar.pack(pady=10)

progress_label = Label(root, text="Waiting for action...", font=("Segoe UI", 11, "bold"))
progress_label.pack(pady=3)

elapsed_label = Label(root, text="⏱️ Elapsed: 0s", font=("Segoe UI", 11), bootstyle="info")
elapsed_label.pack()

eta_label = Label(root, text="⏳ ETA: 0s", font=("Segoe UI", 11), bootstyle="warning")
eta_label.pack()

finish_label = Label(root, text="🕓 Estimated Finish: --:--:--", font=("Segoe UI", 11), bootstyle="success")
finish_label.pack(pady=(0, 10))

# Theme Toggle
darkmode_var = StringVar(value="0")
theme_toggle = Checkbutton(root, text="Dark Mode", variable=darkmode_var, command=toggle_theme, bootstyle="toolbutton")
theme_toggle.pack(pady=5)

log_frame = Frame(root)
log_frame.pack(pady=5, fill="both", expand=True)

scrollbar = Scrollbar(log_frame)
scrollbar.pack(side="right", fill="y")

log_text = Text(log_frame, height=10, wrap="word", yscrollcommand=scrollbar.set, state=DISABLED, font=("Consolas", 10))
log_text.pack(fill="both", expand=True)
scrollbar.config(command=log_text.yview)

root.mainloop()
