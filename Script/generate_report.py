import os
import threading
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from docxtpl import DocxTemplate
import ttkbootstrap as tb
from ttkbootstrap.toast import ToastNotification

# === Toast Helper ===
def show_toast(title, message, duration=3000):
    ToastNotification(title=title, message=message, duration=duration).show_toast()

# === Excel Data Extraction ===
def extract_data_from_excel(file_path, source_key):
    try:
        wb = load_workbook(file_path, data_only=True)
        result_sheet = wb["Result Data"]
        sim_sheet = wb["Simulasi Transaksi"]
        prefix = source_key.lower()[:3]

        return {
            f"{prefix}_ptg_avg": result_sheet["B2"].value,
            f"{prefix}_ptg_min": result_sheet["B4"].value,
            f"{prefix}_trx_min": result_sheet["C4"].value,
            f"{prefix}_trx_avg": result_sheet["C5"].value,
            f"{prefix}_trx_10": sim_sheet["B2"].value,
            f"{prefix}_trx_20": sim_sheet["B3"].value,
            f"{prefix}_trx_30": sim_sheet["B4"].value,
            f"{prefix}_trx_60": sim_sheet["B5"].value,
            f"{prefix}_trx_90": sim_sheet["B6"].value,
            f"{prefix}_trx_120": sim_sheet["B7"].value,
            f"{prefix}_trx_180": sim_sheet["B8"].value,
            f"{prefix}_trx_240": sim_sheet["B9"].value
        }
        
    except Exception:
        show_toast("Error", f"Failed to read: {os.path.basename(file_path)}")
        return {}

# === Word Template Processor ===
def generate_report(template_path, output_path, context):
    tpl = DocxTemplate(template_path)
    tpl.render(context)
    tpl.save(output_path)

# === Main Processing Logic ===
def process_folders():
    parent_dir = folder_path.get()
    template = word_path.get()
    report_maker = name_entry.get()

    if not os.path.isdir(parent_dir) or not os.path.isfile(template) or not report_maker.strip():
        messagebox.showerror("Missing Data", "Please provide all required inputs.")
        return

    branches = [b for b in os.listdir(parent_dir) if os.path.isdir(os.path.join(parent_dir, b))]
    progress_bar['maximum'] = len(branches)
    log_data = []
    preview_box.delete("1.0", "end")

    for idx, branch in enumerate(branches):
        branch_path = os.path.join(parent_dir, branch)
        context = {
            "nama_pembuat": report_maker,
            "nama_cabang": branch,
            "tgl_rilis": datetime.now().strftime("%d %B %Y")
        }
        extracted_data = {}
        preview_box.insert("end", f"[{branch}]\n", "branch")
        found = False

        for folder, prefix in [("eService", "esv"), ("eBranch", "ebr"), ("Star Teller", "str")]:
            result_path = os.path.join(branch_path, folder, "result.xlsx")
            if os.path.isfile(result_path):
                values = extract_data_from_excel(result_path, prefix)
                context.update(values)
                extracted_data.update(values)
                found = True
                preview_box.insert("end", f"  ✓ Found data in {folder}\n", "success")
            else:
                preview_box.insert("end", f"  ✗ Missing {folder}\n", "warning")

        if found:
            preview_box.insert("end", "  Injected data:\n", "info")
            for k, v in extracted_data.items():
                preview_box.insert("end", f"    {k}: {v}\n", "data")

            output_file = os.path.join(branch_path, f"Utilization Report {branch} 2025.docx")
            try:
                generate_report(template, output_file, context)
                preview_box.insert("end", f"  ✔ Report created: {output_file}\n\n", "success")
                log_data.append((branch, "Success", output_file))
            except Exception as e:
                preview_box.insert("end", f"  ✗ Failed to generate report: {e}\n\n", "error")
                log_data.append((branch, f"Error - {e}", ""))
        else:
            preview_box.insert("end", "  ✗ Skipped: No valid Excel files\n\n", "warning")
            log_data.append((branch, "Skipped - No Data", ""))

        progress_bar['value'] = idx + 1
        app.update_idletasks()

    save_log_to_excel(log_data, parent_dir)
    show_toast("Done", "All branches processed.")

def save_log_to_excel(log_data, folder):
    wb = Workbook()
    ws = wb.active
    ws.append(["Branch", "Status", "Output File"])
    for row in log_data:
        ws.append(row)
    log_path = os.path.join(folder, "generation_log.xlsx")
    wb.save(log_path)
    preview_box.insert("end", f"\nLog saved to: {log_path}\n", "info")
    preview_box.see("end")

def threaded_processing():
    threading.Thread(target=process_folders, daemon=True).start()

def center_window(app, width, height):
    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()
    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))
    app.geometry(f"{width}x{height}+{x}+{y}")

# === GUI Setup ===
app = tb.Window(themename="minty")
app.title("Utilization Report Generator 2025")
center_window(app, 850, 700)

tb.Label(app, text="Utilization Report Generator", font=("Helvetica", 18, "bold")).pack(pady=15)

frm = tb.Frame(app, padding=(20, 10))
frm.pack(fill="x", padx=20)

# Row 0: Name Entry
tb.Label(frm, text="Report Maker Name:").grid(row=0, column=0, sticky="w", pady=5)
name_entry = tb.Entry(frm, width=45)
name_entry.grid(row=0, column=1, padx=10)

# Row 1: Folder Selection
tb.Label(frm, text="Select Parent Folder:").grid(row=1, column=0, sticky="w", pady=5)
folder_path = tb.StringVar()
tb.Entry(frm, textvariable=folder_path, width=45).grid(row=1, column=1, padx=10)
tb.Button(frm, text="Browse", command=lambda: folder_path.set(filedialog.askdirectory())).grid(row=1, column=2)

# Row 2: Template File
tb.Label(frm, text="Upload Word Template (.docx):").grid(row=2, column=0, sticky="w", pady=5)
word_path = tb.StringVar()
tb.Entry(frm, textvariable=word_path, width=45).grid(row=2, column=1, padx=10)
tb.Button(frm, text="Browse", command=lambda: word_path.set(filedialog.askopenfilename(filetypes=[("Word files", "*.docx")]))).grid(row=2, column=2)

# Progress Bar
progress_bar = tb.Progressbar(app, length=700, mode='determinate', bootstyle="info-striped")
progress_bar.pack(pady=20)

# Generate Button
tb.Button(app, text="Generate Reports", bootstyle="success", command=threaded_processing).pack(pady=10)

# Preview Panel
preview_frame = tb.LabelFrame(app, text="Preview Panel", padding=10)
preview_frame.pack(fill="both", expand=True, padx=20, pady=10)

preview_box = tb.ScrolledText(preview_frame, height=20, wrap="word", font=("Courier New", 10))
preview_box.pack(fill="both", expand=True)

# Color Tags
preview_box.tag_config("success", foreground="green")
preview_box.tag_config("warning", foreground="orange")
preview_box.tag_config("error", foreground="red")
preview_box.tag_config("info", foreground="blue")
preview_box.tag_config("data", foreground="teal")
preview_box.tag_config("branch", foreground="purple", font=("Courier New", 10, "bold"))

# Footer
footer = tb.Label(app, text="© 2025 Created by SOW 7", font=("Arial", 10), anchor="center")
footer.pack(side="bottom", fill="x", pady=10)

app.mainloop()