import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from ttkbootstrap import Style
from ttkbootstrap.toast import ToastNotification
from datetime import datetime
import locale
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')

class CashDepositApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cash Deposit v1.0.0")
        self.style = Style("flatly")

        w, h = 1200, 850
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        root.geometry(f"{w}x{h}+{x}+{y}")

        self.toast = ToastNotification(title="Welcome", message="Selamat datang di Cash Deposit",
                                       duration=3000, bootstyle="info")
        self.toast.show_toast()

        self.deleted_row = None
        self.build_ui()
        self.update_clock()

    def build_ui(self):
        date_frame = ttk.Frame(self.root, padding=10)
        date_frame.pack(fill='x')
        self.date_label = ttk.Label(date_frame, font=("Helvetica", 12, "bold"))
        self.date_label.pack(anchor='center')

        user_frame = ttk.Labelframe(self.root, text="User Data", padding=10)
        user_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(user_frame, text="User ID:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.user_id_entry = ttk.Entry(user_frame, width=25)
        self.user_id_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(user_frame, text="Seal Number:").grid(row=0, column=2, padx=5, pady=5, sticky='e')
        self.seal_number_entry = ttk.Entry(user_frame, width=25)
        self.seal_number_entry.grid(row=0, column=3, padx=5, pady=5)

        denom_frame = ttk.Labelframe(self.root, text="Table Input Denom", padding=10)
        denom_frame.pack(fill='x', padx=10, pady=10)

        self.denom_fields = {}
        self.denom_labels = ["100", "75", "50", "20", "10", "5", "2", "1", "STAR"]
        for i, label in enumerate(self.denom_labels):
            ttk.Label(denom_frame, text=label).grid(row=0, column=i, padx=5, pady=5)
            entry = ttk.Entry(denom_frame, width=10, justify='right')
            entry.grid(row=1, column=i, padx=5, pady=5)
            entry.bind('<KeyRelease>', self.format_currency_input)
            self.denom_fields[label] = entry

        action_frame = ttk.Frame(self.root)
        action_frame.pack(pady=10)
        ttk.Button(action_frame, text="Submit", command=self.submit_data, width=20).pack(side='left', padx=10)

        self.columns = ["USER-SEAL"] + self.denom_labels + ["Total"]
        self.tree = self.create_table("Data Setoran Kas")
        self.money_tree = self.create_table("Lembar Uang")

        button_frame = ttk.Frame(self.root)
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Delete Selected", command=self.delete_selected_row, width=20).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Clear All", command=self.clear_all, width=20).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Undo Delete", command=self.undo_delete, width=20).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel, width=20).pack(side='left', padx=5)

        footer = ttk.Label(self.root, text="@ Nico Ardian Nugroho SOW 7 - 2025", font=("Helvetica", 9))
        footer.pack(side='bottom', pady=5)

    def create_table(self, title):
        frame = ttk.Labelframe(self.root, text=title, padding=10)
        frame.pack(fill='both', expand=True, padx=10, pady=5)

        tree = ttk.Treeview(frame, columns=self.columns, show='headings', selectmode="browse")
        for col in self.columns:
            tree.heading(col, text=col)
            tree.column(col, anchor='center', width=100)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        return tree

    def update_clock(self):
        days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        now = datetime.now()
        day_name = days[now.weekday()]
        date_string = now.strftime("%d %B %Y")
        time_string = now.strftime("%H.%M.%S WIB")
        self.date_label.config(text=f"{day_name}, {date_string} || {time_string}")
        self.root.after(1000, self.update_clock)

    def format_currency_input(self, event):
        widget = event.widget
        text = widget.get().replace('.', '').replace('Rp', '').replace(',', '').strip()
        if text.isdigit():
            formatted = locale.format_string("%d", int(text), grouping=True)
            widget.delete(0, tk.END)
            widget.insert(0, formatted)

    def parse_currency_input(self, text):
        return int(text.replace('.', '').replace('Rp', '').replace(',', '').strip() or 0)

    def submit_data(self):
        user_id = self.user_id_entry.get().strip()
        seal_number = self.seal_number_entry.get().strip()
        if not user_id or not seal_number:
            self.toast = ToastNotification(title="Input Error", message="User ID and Seal Number cannot be empty", bootstyle="danger")
            self.toast.show_toast()
            return

        user_seal = f"{user_id}/{seal_number}"

        for row_id in self.tree.get_children():
            existing_user_seal = self.tree.item(row_id)['values'][0]
            try:
                existing_user_id, existing_seal = existing_user_seal.split('/')
            except Exception:
                continue
            if seal_number == existing_seal:
                if user_id.startswith(existing_user_id) or existing_user_id.startswith(user_id):
                    self.toast = ToastNotification(
                        title="Duplicate Error",
                        message=f"USER-SEAL '{user_seal}' conflicts with existing '{existing_user_seal}'.",
                        bootstyle="danger")
                    self.toast.show_toast()
                    return

        try:
            values = [self.parse_currency_input(self.denom_fields[key].get()) for key in self.denom_fields]
        except ValueError:
            self.toast = ToastNotification(title="Input Error", message="Please enter valid numbers", bootstyle="danger")
            self.toast.show_toast()
            return

        total_rupiah = sum(values)
        formatted_values = [locale.currency(v, grouping=True) for v in values]
        formatted_total = locale.currency(total_rupiah, grouping=True)
        self.tree.insert("", "end", values=[user_seal] + formatted_values + [formatted_total])

        converted = []
        for k, v in zip(self.denom_labels, values):
            if k == "STAR":
                converted.append(locale.currency(v, grouping=True))
            else:
                denom = int(k) * 1000
                lembar = v // denom
                converted.append(str(lembar))

        self.money_tree.insert("", "end", values=[user_seal] + converted + [formatted_total])

        self.toast = ToastNotification(title="Success", message="Data successfully added", bootstyle="success")
        self.toast.show_toast()

    def delete_selected_row(self):
        selected = self.tree.selection()
        if selected:
            if messagebox.askyesno("Confirm", "Are you sure you want to delete the selected row?"):
                index = self.tree.index(selected[0])
                self.deleted_row = (
                    index,
                    self.tree.item(selected[0])['values'],
                    self.money_tree.item(self.money_tree.get_children()[index])['values']
                )
                self.tree.delete(selected[0])
                self.money_tree.delete(self.money_tree.get_children()[index])
                self.toast = ToastNotification(title="Deleted", message="Row has been deleted.", bootstyle="warning")
                self.toast.show_toast()

    def undo_delete(self):
        if self.deleted_row:
            index, values1, values2 = self.deleted_row
            self.tree.insert("", index, values=values1)
            self.money_tree.insert("", index, values=values2)
            self.deleted_row = None
            self.toast = ToastNotification(title="Undo", message="Last deletion has been undone.", bootstyle="info")
            self.toast.show_toast()

    def clear_all(self):
        if not messagebox.askyesno("Confirm", "Are you sure you want to clear all inputs and table data?"):
            return

        self.user_id_entry.delete(0, tk.END)
        self.seal_number_entry.delete(0, tk.END)
        for entry in self.denom_fields.values():
            entry.delete(0, tk.END)
        for row in self.tree.get_children():
            self.tree.delete(row)
        for row in self.money_tree.get_children():
            self.money_tree.delete(row)
        self.toast = ToastNotification(title="Cleared", message="All data has been cleared.", bootstyle="danger")
        self.toast.show_toast()

    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data Setoran Kas"
        ws2 = wb.create_sheet("Lembar Uang")

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)

        def write_sheet(tree, ws):
            for col_num, col_name in enumerate(self.columns, start=1):
                cell = ws.cell(row=1, column=col_num, value=col_name)
                cell.font = bold_font
                cell.border = thin_border
            for row_num, row_id in enumerate(tree.get_children(), start=2):
                values = tree.item(row_id)['values']
                for col_num, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

        write_sheet(self.tree, ws1)
        write_sheet(self.money_tree, ws2)

        try:
            wb.save(file_path)
            messagebox.showinfo("Success", f"Data exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")

if __name__ == '__main__':
    root = tk.Tk()
    app = CashDepositApp(root)
    root.mainloop()
